"""
Artistic Dental Studio — Retention Logic
=========================================
Python port of Lexie's v3 retention dashboard logic (HTML/JS → Python).

Public API
----------
parse_retention_period(path)   -> dict {title, months[3], doctors[]}
parse_case_file(path)          -> dict {customer_id: CaseEntry}
load_case_history(path)        -> dict (reads cache/case_history.csv)
save_case_history(map, path)   -> writes cache/case_history.csv
merge_case_maps(existing, fresh) -> dict
compute_retention(periods, case_map) -> tuple of DataFrames
    (periods_df, doctors_df, master_df)

Retention rules (from Lexie's getStatus):
    retained  : sent ≥1 REAL case in M3 (third month of window)
    at_risk   : real case in M2, none in M3
    lost      : no real cases in M2 or M3
    REAL case = NOT (remake-100 OR adjustment-discount)

Window scope (from Lexie's getFullStatus):
    3M  window: M1, M2, M3 (period months)
    6M  window: M4, M5, M6 (months following the 3M window)
    12M window: M7..M12   (next six months after 6M)
"""

from __future__ import annotations

import re
import csv
import json
import logging
import subprocess
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd

log = logging.getLogger("retention")

# ── Constants (matches Lexie's MN/MNF/PFXMAP) ──────────────────────────────────
MONTH_ABBR = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTH_PFX = {a.lower(): i for i, a in enumerate(MONTH_ABBR)}

# Active marker regex — matches Lexie's /^[Xx√✓vV1]$/
ACTIVE_RE = re.compile(r'^[XxvV1√✓]$')

# Words that look like data but are actually summary/header rows
HEADER_WORDS = {'customerid', 'custid', 'name', 'status',
                'digital', 'city', 'product', 'type'}

# Case# is always 6 digits in Magic Touch
CASE_NUM_RE = re.compile(r'^\d{6}$')

# Adjustment / Remake-100 detection (from Lexie):
#   isRemake100 = (/100/.test(remake) && /remake/i.test(remake)) || /adjustment/i.test(remake)
def _is_non_billable(remake_text: str) -> bool:
    s = (remake_text or '').strip()
    if not s:
        return False
    if 'adjustment' in s.lower():
        return True
    if '100' in s and re.search(r'remake', s, re.IGNORECASE):
        return True
    return False


# ══════════════════════════════════════════════════════════════════════════════
#  FILE READING — handles .xls (via libreoffice), .xlsx (openpyxl), .csv
# ══════════════════════════════════════════════════════════════════════════════

def _read_sheet_as_rows(path: Path) -> list[list]:
    """
    Read first sheet of an .xls/.xlsx/.csv as a list of rows (lists of cell values).
    Strings are kept as strings; date cells become Python date/datetime if available.
    """
    suffix = path.suffix.lower()

    if suffix == '.csv':
        return _read_csv_rows(path)

    if suffix == '.xlsx':
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [[c if c is not None else '' for c in row]
                    for row in ws.iter_rows(values_only=True)]
            wb.close()
            return rows
        except Exception as exc:
            log.warning("openpyxl failed on %s, falling back to libreoffice: %s",
                        path.name, exc)

    if suffix == '.xls':
        # Try xlrd first (if installed and supports .xls — xlrd 1.2.0)
        try:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            sh = wb.sheet_by_index(0)
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == 3:  # XL_CELL_DATE
                        row.append(xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode))
                    else:
                        row.append(cell.value)
                rows.append(row)
            return rows
        except Exception as exc:
            log.info("xlrd unavailable for %s, using libreoffice: %s", path.name, exc)

    # Fallback: convert via libreoffice headless and read CSV
    return _read_via_libreoffice(path)


def _read_csv_rows(path: Path) -> list[list]:
    # latin-1 because Magic Touch case exports use it (matches pipeline.py convention)
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(path, 'r', encoding=enc, newline='') as f:
                return [row for row in csv.reader(f)]
        except UnicodeDecodeError:
            continue
    raise IOError(f"Could not decode {path}")


def _find_libreoffice() -> str | None:
    """Locate the LibreOffice headless executable across Mac / Linux / Windows."""
    import shutil, os
    # On PATH?
    for name in ('libreoffice', 'soffice', 'soffice.exe'):
        p = shutil.which(name)
        if p:
            return p
    # Common Windows install paths
    candidates = [
        r'C:\Program Files\LibreOffice\program\soffice.exe',
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
        # macOS
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


def _read_via_libreoffice(path: Path) -> list[list]:
    """Convert via libreoffice --headless and read the resulting CSV."""
    exe = _find_libreoffice()
    if not exe:
        raise IOError(
            f"Cannot read {path.name}: LibreOffice not found on this system. "
            "Install LibreOffice (https://www.libreoffice.org/download/), "
            "OR install xlrd 1.2.0 with `py -m pip install xlrd==1.2.0`, "
            "OR re-export your file as .xlsx or .csv from Magic Touch."
        )
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        subprocess.run(
            [exe, '--headless', '--convert-to', 'csv',
             '--outdir', str(tmp_path), str(path)],
            check=True, capture_output=True, timeout=120,
        )
        out_csv = tmp_path / (path.stem + '.csv')
        if not out_csv.exists():
            # libreoffice strips quotes/spaces from filenames sometimes
            csvs = list(tmp_path.glob('*.csv'))
            if not csvs:
                raise IOError(f"LibreOffice produced no output for {path.name}")
            out_csv = csvs[0]
        return _read_csv_rows(out_csv)


# ══════════════════════════════════════════════════════════════════════════════
#  PARSE RETENTION PERIOD — port of Lexie's parseRet
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class RetentionDoctor:
    customer_id: str
    name: str
    city: str = ''
    type: str = ''            # 'New' or 'Returning'
    digital: bool = False
    first_product: str = ''
    m1_active_xls: bool = False  # checkmark in M1 column of XLS
    m2_active_xls: bool = False
    m3_active_xls: bool = False


def parse_retention_period(path: str | Path) -> dict:
    """
    Parse a single retention-period XLS/XLSX/CSV file.
    Returns: {title, months: [{month,year}, ...], doctors: [RetentionDoctor]}
    Returns None if the file isn't a standard retention grid (e.g., May '26 layout).
    """
    p = Path(path)
    aoa = _read_sheet_as_rows(p)
    if not aoa:
        log.warning("Empty file: %s", p.name)
        return None

    # Auto-detect header row by scanning first 8 rows for "customerid" or "custid"
    header_row_idx, title_row_idx = 1, 0
    for i in range(min(8, len(aoa))):
        joined = '|'.join(str(c) for c in (aoa[i] or [])).lower()
        if 'customerid' in joined or 'custid' in joined:
            header_row_idx = i
            break
    else:
        # No header found — may be the May '26 different-format file
        log.warning("Could not find CustomerID header in %s — skipping (likely a non-standard layout).",
                    p.name)
        return None

    # Title row = first non-empty row above header
    for i in range(header_row_idx - 1, -1, -1):
        val = ''
        if aoa[i]:
            val = str(aoa[i][0] if aoa[i][0] else (aoa[i][1] if len(aoa[i]) > 1 else '')).strip()
        if val and val.lower() != 'nan':
            title_row_idx = i
            break

    title_row = aoa[title_row_idx] or []
    title = str(title_row[0] if title_row and title_row[0]
                else (title_row[1] if len(title_row) > 1 else '')).strip()

    header_row = aoa[header_row_idx] or []
    data_start = header_row_idx + 1

    # Column auto-detection
    digital_col, status_col, custid_col, name_col = 0, 1, 2, 3
    city_col = -1
    first_product_col = -1
    month_cols: list[tuple[int, int]] = []  # (col_idx, month_idx 0-11)

    for j, h in enumerate(header_row):
        raw = str(h or '')
        hl = re.sub(r'[\s*.]', '', raw).lower()
        if 'customerid' in hl or 'custid' in hl:
            custid_col = j
        elif hl == 'status':
            status_col = j
        elif hl.startswith('dig'):
            digital_col = j
        elif re.search(r'product|casenumber|1stcase|firstcase|1stprod', hl):
            first_product_col = j
        elif hl == 'name':
            name_col = j
        elif hl in ('city', 'location'):
            city_col = j
        else:
            # Try to match a month — supports several header formats:
            #   "Jan", "January"           → abbreviation prefix
            #   "2025-01", "2025-01-15"    → ISO
            #   "01/25/2025", "1/25/26"    → US MM/DD/YYYY
            mi = -1
            iso = re.search(r'\b(\d{4})-(\d{2})', raw)
            us = re.search(r'\b(\d{1,2})/\d{1,2}/\d{2,4}\b', raw)
            if iso:
                mi = int(iso.group(2)) - 1
            elif us:
                mi = int(us.group(1)) - 1
            else:
                for k, abbr in enumerate(MONTH_ABBR):
                    if hl.startswith(abbr.lower()):
                        mi = k
                        break
            if 0 <= mi < 12:
                month_cols.append((j, mi))

    # If no city column found, use the first unused column to the right of custid
    if city_col == -1:
        used = {digital_col, status_col, custid_col, name_col}
        if first_product_col >= 0:
            used.add(first_product_col)
        used.update(c for c, _ in month_cols)
        for j in range(len(header_row)):
            if j not in used and j > custid_col:
                city_col = j
                break

    # Bail out on non-standard layouts (e.g., May '26 case-detail report).
    # A real retention grid has a monthly title AND at least 3 month columns.
    if len(month_cols) < 3:
        log.warning("Skipping %s — only %d month column(s) detected (not a standard retention grid)",
                    p.name, len(month_cols))
        return None
    if re.fullmatch(r'\d{1,2}/\d{1,2}/\d{2,4}', title.strip()):
        log.warning("Skipping %s — title is a raw date (%s), not a retention period",
                    p.name, title)
        return None

    months = _title_to_months(title, month_cols, filename=p.name)

    # The last 3 month columns are M1/M2/M3 of the window (matches Lexie)
    last_three = month_cols[-3:] if len(month_cols) >= 3 else month_cols
    m_cols = [c for c, _ in last_three] + [None, None, None]
    m1c, m2c, m3c = m_cols[0], m_cols[1], m_cols[2]

    doctors: list[RetentionDoctor] = []
    for i in range(data_start, len(aoa)):
        r = aoa[i] or []
        cid = str(_cell(r, custid_col)).strip()
        if not cid:
            continue
        # Skip numeric-only rows, summary rows, header words, short tokens
        if re.fullmatch(r'\d+\.?\d*', cid):
            continue
        if re.match(r'^(new|returning|total)\s*[=:]', cid, re.IGNORECASE):
            continue
        if len(cid) < 3 or cid.lower() == 'nan':
            continue
        if cid.lower() in HEADER_WORDS:
            continue

        doctors.append(RetentionDoctor(
            customer_id=cid,
            name=str(_cell(r, name_col)).strip(),
            city=str(_cell(r, city_col)).strip() if city_col >= 0 else '',
            type=str(_cell(r, status_col)).strip(),
            digital=str(_cell(r, digital_col)).strip() == '*',
            first_product=str(_cell(r, first_product_col)).strip() if first_product_col >= 0 else '',
            m1_active_xls=_is_active(_cell(r, m1c)),
            m2_active_xls=_is_active(_cell(r, m2c)),
            m3_active_xls=_is_active(_cell(r, m3c)),
        ))

    return {
        'title': title,
        'months': months,                 # list of {'month':int, 'year':int} — index 0 = M1, 2 = M3
        'doctors': doctors,
        'source_file': p.name,
    }


def _cell(row: list, idx: Optional[int]):
    if idx is None or idx < 0 or idx >= len(row):
        return ''
    return row[idx] if row[idx] is not None else ''


def _is_active(value) -> bool:
    s = str(value or '').strip()
    return bool(ACTIVE_RE.match(s))


def _title_to_months(title: str,
                     header_month_cols: list[tuple[int, int]],
                     filename: str = '') -> list[dict]:
    """
    Extract the 3-month window (M1/M2/M3) from a retention period title.
    Falls back to filename for the year (e.g. "April '25 Retention.xls" → 2025)
    when the title doesn't include one.
    """
    clean = re.sub(r'retention|doctors?', '', title, flags=re.IGNORECASE).lower()
    clean = re.sub(r'\s+', ' ', clean).strip()
    parts = clean.split(' ')

    mi = -1
    for p in parts:
        if not p or len(p) < 3:
            continue
        pfx = p[:3]
        if pfx in MONTH_PFX:
            mi = MONTH_PFX[pfx]
            break
        # Fuzzy: catches "septermber" → match by 3-char prefix overlap
        for key, idx in MONTH_PFX.items():
            if p.startswith(key[:3]) or key.startswith(p[:3]):
                mi = idx
                break
        if mi != -1:
            break

    yr = None
    for p in parts:
        try:
            n = int(p)
            if 2020 <= n <= 2035:
                yr = n
                break
        except ValueError:
            pass

    # Year fallback #1: pull from filename — handles "April '25 Retention.xls" etc.
    if yr is None and filename:
        m4 = re.search(r"\b(20\d{2})\b", filename)
        if m4:
            yr = int(m4.group(1))
        else:
            m2 = re.search(r"'(\d{2})\b", filename)
            if m2:
                yr = 2000 + int(m2.group(1))

    # Month fallback: use last header month column if title didn't yield one
    if mi == -1 and header_month_cols:
        mi = header_month_cols[-1][1]

    if mi == -1:
        return []

    if yr is None:
        # Last resort — assume current calendar year
        yr = date.today().year

    # Period covers months (mi-2, mi-1, mi)
    out = []
    for offset in range(3):
        m = mi - (2 - offset)
        y = yr
        while m < 0:
            m += 12
            y -= 1
        out.append({'month': m, 'year': y})
    return out


# ══════════════════════════════════════════════════════════════════════════════
#  PARSE CASE FILE — port of Lexie's parseCases / parseCSVCases
# ══════════════════════════════════════════════════════════════════════════════

# Magic Touch case-list column indices (from the briefing, confirmed against real file)
CASE_COL = {
    'case_number':     1,
    'account_id':      4,
    'doctor_name':     6,
    'discount':        7,
    'discount_amount': 11,
    'pan_number':      14,
    'patient_name':    16,
    'date_in':         22,
    'ship_date':       27,
    'due_date':        29,
    'remake':          31,
}


@dataclass
class CaseRow:
    case_number: str
    account_id: str
    doctor_name: str
    date_in: date
    discount: str
    remake: str
    pan_number: str = ''
    patient_name: str = ''
    ship_date: Optional[date] = None
    due_date: Optional[date] = None
    is_non_billable: bool = False


def parse_case_file(path: str | Path) -> list[CaseRow]:
    """
    Parse a Magic Touch Case List by Date export (.xls/.xlsx/.csv).
    Returns deduplicated list of CaseRow (dedup key = case_number).
    """
    p = Path(path)
    aoa = _read_sheet_as_rows(p)
    if not aoa:
        return []

    out: list[CaseRow] = []
    seen: set[str] = set()

    def _coerce_str(v):
        """Coerce cell to string, stripping the .0 that xlrd adds to numeric cells."""
        if v is None:
            return ''
        if isinstance(v, (int,)):
            return str(v)
        if isinstance(v, float):
            if v.is_integer():
                return str(int(v))
            return str(v)
        return str(v).strip()

    for row in aoa:
        if not row:
            continue
        case_num = _coerce_str(_cell(row, CASE_COL['case_number']))
        if not CASE_NUM_RE.match(case_num):
            continue  # skip header/summary rows — only real 6-digit case# rows
        if case_num in seen:
            continue
        seen.add(case_num)

        account_id = _coerce_str(_cell(row, CASE_COL['account_id']))
        if not account_id or len(account_id) < 3:
            continue

        date_in = _coerce_date(_cell(row, CASE_COL['date_in']))
        if date_in is None:
            continue  # case with no date is unusable for retention

        remake_field = str(_cell(row, CASE_COL['remake'])).strip()
        out.append(CaseRow(
            case_number=case_num,
            account_id=account_id,
            doctor_name=str(_cell(row, CASE_COL['doctor_name'])).strip(),
            date_in=date_in,
            discount=str(_cell(row, CASE_COL['discount'])).strip(),
            remake=remake_field,
            pan_number=str(_cell(row, CASE_COL['pan_number'])).strip(),
            patient_name=str(_cell(row, CASE_COL['patient_name'])).strip(),
            ship_date=_coerce_date(_cell(row, CASE_COL['ship_date'])),
            due_date=_coerce_date(_cell(row, CASE_COL['due_date'])),
            is_non_billable=_is_non_billable(remake_field),
        ))

    return out


def _coerce_date(val) -> Optional[date]:
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y',
                '%m-%d-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # pandas as last resort (handles many odd formats)
    try:
        ts = pd.to_datetime(s, errors='coerce')
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  CASE HISTORY — long-lived store, merged across uploads
# ══════════════════════════════════════════════════════════════════════════════

CASE_HISTORY_COLS = [
    'case_number', 'account_id', 'doctor_name', 'date_in',
    'discount', 'remake', 'pan_number', 'patient_name',
    'ship_date', 'due_date', 'is_non_billable',
]


def cases_to_dataframe(cases: list[CaseRow]) -> pd.DataFrame:
    if not cases:
        return pd.DataFrame(columns=CASE_HISTORY_COLS)
    rows = []
    for c in cases:
        rows.append({
            'case_number': c.case_number,
            'account_id': c.account_id,
            'doctor_name': c.doctor_name,
            'date_in': c.date_in.isoformat() if c.date_in else '',
            'discount': c.discount,
            'remake': c.remake,
            'pan_number': c.pan_number,
            'patient_name': c.patient_name,
            'ship_date': c.ship_date.isoformat() if c.ship_date else '',
            'due_date': c.due_date.isoformat() if c.due_date else '',
            'is_non_billable': bool(c.is_non_billable),
        })
    return pd.DataFrame(rows, columns=CASE_HISTORY_COLS)


# Dummy / test accounts excluded from every metric across the portal.
# Keep in sync with pipeline.py and pipeline_logistics.py.
EXCLUDED_ACCOUNT_IDS = {"LAWMUR"}


def load_case_history(path: str | Path) -> pd.DataFrame:
    """Load the long-lived case_history.csv, or an empty frame if missing.

    Runtime safety net: filters out dummy/test accounts even if any historical
    rows remain in the persisted file.
    """
    p = Path(path)
    if not p.exists():
        return pd.DataFrame(columns=CASE_HISTORY_COLS)
    df = pd.read_csv(p, dtype={'case_number': str, 'account_id': str},
                     keep_default_na=False)
    if not df.empty and 'account_id' in df.columns:
        norm = df['account_id'].fillna('').astype(str).str.strip().str.upper()
        df = df[~norm.isin(EXCLUDED_ACCOUNT_IDS)].copy()
    return df


def save_case_history(df: pd.DataFrame, path: str | Path) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(p, index=False)


def merge_case_dataframes(existing: pd.DataFrame, fresh: pd.DataFrame) -> pd.DataFrame:
    """
    Merge two case dataframes by case_number (last-write-wins for matching keys).
    Returns merged frame sorted by date_in descending.
    """
    if existing.empty:
        return fresh.copy()
    if fresh.empty:
        return existing.copy()
    combined = pd.concat([existing, fresh], ignore_index=True)
    # Keep the LAST occurrence (so a fresh row overwrites a stale one)
    combined = combined.drop_duplicates(subset=['case_number'], keep='last')
    combined = combined.sort_values('date_in', ascending=False)
    return combined.reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
#  STATUS COMPUTATION — port of Lexie's getStatus / getFullStatus
# ══════════════════════════════════════════════════════════════════════════════

def _build_doctor_month_set(history: pd.DataFrame) -> dict[str, set[tuple[int, int]]]:
    """
    Build {customer_id: {(month, year), ...}} of months in which the doctor sent
    at least one BILLABLE (real) case. Skips remake-100 / adjustment cases.
    """
    if history.empty:
        return {}
    df = history[history['is_non_billable'] != True].copy()  # noqa: E712
    df['date_in'] = pd.to_datetime(df['date_in'], errors='coerce')
    df = df.dropna(subset=['date_in'])
    df['__m'] = df['date_in'].dt.month - 1  # 0-indexed to match Lexie
    df['__y'] = df['date_in'].dt.year
    out: dict[str, set[tuple[int, int]]] = {}
    for cid, grp in df.groupby('account_id'):
        out[str(cid)] = set(zip(grp['__m'].astype(int), grp['__y'].astype(int)))
    return out


def _build_doctor_last_case(history: pd.DataFrame) -> dict[str, str]:
    """Most-recent date_in per doctor (ISO string), considering REAL cases only."""
    if history.empty:
        return {}
    df = history[history['is_non_billable'] != True].copy()  # noqa: E712
    df['date_in'] = pd.to_datetime(df['date_in'], errors='coerce')
    df = df.dropna(subset=['date_in'])
    if df.empty:
        return {}
    last = df.groupby('account_id')['date_in'].max()
    return {str(cid): d.date().isoformat() for cid, d in last.items()}


def _window_months(m1: dict, start_offset: int, count: int) -> list[tuple[int, int]]:
    """Build (month, year) tuples starting at m1+start_offset, length=count."""
    out = []
    m = m1['month'] + start_offset
    y = m1['year']
    for _ in range(count):
        while m > 11:
            m -= 12
            y += 1
        out.append((m, y))
        m += 1
    return out


def compute_retention(periods: list[dict],
                      case_history: pd.DataFrame
                      ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Compute retention status for each (period, doctor) pair.

    Returns
    -------
    periods_df : DataFrame
        One row per loaded period: id, title, year, m1_label, m2_label, m3_label,
        total, retained, at_risk, lost, retained_pct, source_file.

    doctors_df : DataFrame
        One row per (period, doctor): period_id, customer_id, name, city, type,
        digital, m1_active, m2_active, m3_active, status_3m, status_6m,
        status_12m, m3_remake_only_flag, last_case_date, source_file.

    master_df : DataFrame
        One row per unique doctor across all periods, with the most recent
        period's status and last_case_date.
    """
    doc_months = _build_doctor_month_set(case_history)
    doc_last_case = _build_doctor_last_case(case_history)
    have_cases = len(doc_months) > 0

    # Build month-cases map for "remake only" detection (per Lexie's m3HasRemakeOnly)
    monthcases_per_doc: dict[str, set[tuple[int, int]]] = {}
    if not case_history.empty:
        df_all = case_history.copy()
        df_all['date_in'] = pd.to_datetime(df_all['date_in'], errors='coerce')
        df_all = df_all.dropna(subset=['date_in'])
        df_all['__m'] = df_all['date_in'].dt.month - 1
        df_all['__y'] = df_all['date_in'].dt.year
        for cid, grp in df_all.groupby('account_id'):
            monthcases_per_doc[str(cid)] = set(zip(grp['__m'].astype(int),
                                                   grp['__y'].astype(int)))

    period_records = []
    doctor_records = []

    # Sort periods chronologically by M3 (month index 2)
    def _sort_key(p):
        m3 = p['months'][2] if len(p.get('months', [])) >= 3 else None
        return (m3['year'], m3['month']) if m3 else (0, 0)

    sorted_periods = sorted(periods, key=_sort_key)

    for pi, period in enumerate(sorted_periods):
        if len(period.get('months', [])) < 3:
            log.warning("Period '%s' has no resolvable months — skipped",
                        period.get('title'))
            continue
        period_id = f"p{pi+1:02d}"
        m1 = period['months'][0]
        m2 = period['months'][1]
        m3 = period['months'][2]
        m6_window = _window_months(m1, 3, 3)   # M4–M6
        m12_window = _window_months(m1, 6, 6)  # M7–M12

        ret = ari = lost = 0
        for d in period['doctors']:
            cid = d.customer_id
            if have_cases and cid in doc_months:
                months_for_doc = doc_months[cid]
                m1_active = (m1['month'], m1['year']) in months_for_doc
                m2_active = (m2['month'], m2['year']) in months_for_doc
                m3_active = (m3['month'], m3['year']) in months_for_doc
            else:
                m1_active = d.m1_active_xls
                m2_active = d.m2_active_xls
                m3_active = d.m3_active_xls

            if m3_active:
                status_3m = 'retained'
            elif m2_active:
                status_3m = 'at_risk'
            else:
                status_3m = 'lost'

            # 6M / 12M (Lexie: any real case in window => retained, else lost)
            if have_cases and cid in doc_months:
                status_6m = 'retained' if any(mk in doc_months[cid] for mk in m6_window) else 'lost'
                status_12m = 'retained' if any(mk in doc_months[cid] for mk in m12_window) else 'lost'
            else:
                status_6m = ''
                status_12m = ''

            # m3HasRemakeOnly: month 3 had ANY cases but no REAL cases
            m3_key = (m3['month'], m3['year'])
            m3_remake_only = (
                cid in monthcases_per_doc
                and m3_key in monthcases_per_doc[cid]
                and not m3_active
            )

            if status_3m == 'retained':
                ret += 1
            elif status_3m == 'at_risk':
                ari += 1
            else:
                lost += 1

            doctor_records.append({
                'period_id': period_id,
                'customer_id': cid,
                'name': d.name,
                'city': d.city,
                'type': d.type,
                'digital': d.digital,
                'first_product': d.first_product,
                'm1_active': m1_active,
                'm2_active': m2_active,
                'm3_active': m3_active,
                'status_3m': status_3m,
                'status_6m': status_6m,
                'status_12m': status_12m,
                'm3_remake_only': m3_remake_only,
                'last_case_date': doc_last_case.get(cid, ''),
                'source_file': period.get('source_file', ''),
            })

        total = ret + ari + lost
        period_records.append({
            'period_id': period_id,
            'title': period['title'],
            'year': m3['year'],
            'm1_label': f"{MONTH_ABBR[m1['month']]} {m1['year']}",
            'm2_label': f"{MONTH_ABBR[m2['month']]} {m2['year']}",
            'm3_label': f"{MONTH_ABBR[m3['month']]} {m3['year']}",
            'total': total,
            'retained': ret,
            'at_risk': ari,
            'lost': lost,
            'retained_pct': round(ret / total * 100, 1) if total else 0.0,
            'source_file': period.get('source_file', ''),
        })

    periods_df = pd.DataFrame(period_records)
    doctors_df = pd.DataFrame(doctor_records)

    # Master = most recent period entry per doctor
    if not doctors_df.empty:
        # Join with periods_df to pick up year/m3_label for chronological sorting
        merged = doctors_df.merge(periods_df[['period_id', 'year', 'm3_label']],
                                  on='period_id', how='left')
        merged['__sort'] = merged['period_id']  # period_ids are zero-padded chronological
        master_df = (merged.sort_values('__sort')
                          .groupby('customer_id', as_index=False)
                          .last()
                          .drop(columns='__sort'))
        master_df = master_df.sort_values('name')
    else:
        master_df = pd.DataFrame()

    return periods_df, doctors_df, master_df




def compute_doctor_activity(case_history: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate per-doctor monthly activity from the case history.
    Returns a clean PHI-free DataFrame: [account_id, month, year, status]
    where status is 'real' (sent >=1 billable case in that month) or
    'remake' (only non-billable cases). Safe to commit to GitHub.
    """
    if case_history.empty:
        return pd.DataFrame(columns=['account_id', 'month', 'year', 'status'])
    df = case_history.copy()
    df['date_in'] = pd.to_datetime(df['date_in'], errors='coerce')
    df = df.dropna(subset=['date_in'])
    df['month'] = df['date_in'].dt.month - 1   # 0-indexed (matches Lexie)
    df['year']  = df['date_in'].dt.year
    df['billable'] = ~df['is_non_billable'].astype(str).str.lower().isin(('true', '1'))
    rows = []
    for (cid, m, y), grp in df.groupby(['account_id', 'month', 'year']):
        rows.append({
            'account_id': str(cid),
            'month': int(m),
            'year':  int(y),
            'status': 'real' if grp['billable'].any() else 'remake',
        })
    return pd.DataFrame(rows).sort_values(['account_id', 'year', 'month'])




def compute_case_summary(case_history: pd.DataFrame) -> pd.DataFrame:
    """
    Sanitized per-case summary for the doctor detail modal. Keeps only the
    columns that are safe to commit to GitHub (no patient names, no pan #s).
    Returns: [account_id, case_number, date_in, doctor_name, is_non_billable]
    """
    if case_history.empty:
        return pd.DataFrame(columns=['account_id', 'case_number', 'date_in',
                                     'doctor_name', 'is_non_billable'])
    keep = ['account_id', 'case_number', 'date_in', 'doctor_name', 'is_non_billable']
    df = case_history[[c for c in keep if c in case_history.columns]].copy()
    df = df.dropna(subset=['date_in'])
    return df.sort_values('date_in', ascending=False)


# -----------------------------------------------------------------------------
#  Convenience: load all periods from a folder
# -----------------------------------------------------------------------------

def load_all_periods(folder, pattern: str = '*Retention*') -> list[dict]:
    """Parse every retention-period file in a folder. Skips non-standard layouts."""
    folder = Path(folder)
    out: list[dict] = []
    seen: set[str] = set()
    for ext in ('.xls', '.xlsx', '.csv'):
        for p in sorted(folder.glob(f'{pattern}{ext}')):
            if p.name in seen:
                continue
            seen.add(p.name)
            try:
                parsed = parse_retention_period(p)
                if parsed and parsed['doctors']:
                    out.append(parsed)
                    log.info("Loaded period: %s (%d doctors)",
                             parsed['title'] or p.name, len(parsed['doctors']))
                else:
                    log.warning("Skipped (no data or non-standard): %s", p.name)
            except Exception as exc:
                log.error("Failed to parse %s: %s", p.name, exc)
    return out
