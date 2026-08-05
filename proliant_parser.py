"""
proliant_parser.py
==================
Parses the Proliant "Payroll Register" export (single-period or YTD) into
per-employee, per-pay-period gross wages, mapped onto our dashboard areas.

PRIVACY-SENSITIVE: payroll dollars are per-person compensation. Parsed output
goes ONLY into the local SQLite DB (labor_history, source='actual') — never
into cache/latest/ (which is committed and served to the cloud app).

Format notes (learned from the real exports):
  * Sheet "Table 1", a PRINT-LAYOUT dump ~66-78 columns wide. Each employee ×
    pay period is a ~7-row block. Two layouts appear in the SAME file: values
    spread across cells ("Emp Id" | "310") and values baked into one wrapped
    text cell ("Emp Id     ALMSEN"). Both are handled.
  * Employees are grouped under "Department: (code)Name" section headers, and
    an employee genuinely split across departments appears under each one.
  * Emp Id is ALPHANUMERIC — mostly numeric ("310") but sometimes a
    MagicTouch-style code ("ALMSEN"). Parsing it as digits-only silently
    shatters one person into several fake employees.
  * A block ends at "Total Earnings <hours> <gross>". One employee can have
    MORE THAN ONE block per check date (e.g. a separate Type: Bonus check) —
    those are additional real dollars and are summed, not de-duplicated.
  * Pay is SEMI-MONTHLY (24/yr). The check date runs a few days after the
    period it pays: day >= 15 pays the 1st-15th of that month; day <= 10 pays
    the 16th-end of the PREVIOUS month.
"""
from __future__ import annotations

import calendar
import re
from datetime import date, timedelta
from pathlib import Path

import pandas as pd

# Proliant department code → how its payroll is attributed, as a list of
# (dashboard, area, weight) with weights summing to 1.0. Areas match the
# technicians table so actuals line up with the dashboards' area drill-down.
# None = non-production overhead, dropped from labor attribution.
#
# Per Danny (2026-08-05): on payroll, "Crown & Bridge" is synonymous with Fixed,
# and ALL Model/Die work splits 50/50 between Fixed and Removable (it feeds both).
PROLIANT_DEPT_MAP: dict[str, list[tuple[str, str, float]] | None] = {
    "101": None,                                    # Distribution (shipping)
    "102": [("Fixed", "Model/Die", 0.5),            # Model/Die serves both
            ("Removable", "Model/Die", 0.5)],
    "107": [("Fixed", "Crown & Bridge", 1.0)],
    "108": [("Fixed", "CAD/CAM", 1.0)],
    "113": [("Fixed", "Ceramics", 1.0)],
    "117": [("Removable", "Removables", 1.0)],
    "118": [("Removable", "Chairside", 1.0)],
    "180": None,                                    # Selling
    "181": None,                                    # Sales & Marketing
    "190": None,                                    # Overhead
    "191": None,                                    # Clinical Technical Services
    "195": None,                                    # Administration
}

# Per-person attribution that overrides the department mapping, keyed by
# normalized payroll name. These are people whose Proliant department does not
# reflect the work they actually do — either a genuine cross-department split or
# a clock-in that lands in the wrong department. Danny's calls, 2026-08-05:
PAYROLL_ATTRIBUTION_OVERRIDES: dict[str, list[tuple[str, str, float]]] = {
    # Runs Surgical Guides (Removable) and CAD/CAM (Fixed) — split for now.
    "gorbach|matthew":  [("Fixed", "CAD/CAM", 0.5),
                         ("Removable", "Removables", 0.5)],
    # Payroll says Crown & Bridge; he is Fixed.
    "deleon|alvin":     [("Fixed", "Crown & Bridge", 1.0)],
    # Actually logistics, which serves both departments.
    "ciaccio|jianni":   [("Fixed", "Logistics", 0.5),
                         ("Removable", "Logistics", 0.5)],
    # Clocks into CAD/CAM but the work is Model/Die → the 50/50 Model/Die rule.
    "flores|brenda":    [("Fixed", "Model/Die", 0.5),
                         ("Removable", "Model/Die", 0.5)],
    # Payroll says Model/Die; he is Fixed 100%.
    "arevalo|henry":    [("Fixed", "Crown & Bridge", 1.0)],
    # Payroll says Model/Die; he runs printing → CAD/CAM.
    "o'hale|john":      [("Fixed", "CAD/CAM", 1.0)],
    # Punches Model/Die and never switches; the work is Ceramics.
    "salazar|paloma":   [("Fixed", "Ceramics", 1.0)],
}


def resolve_targets(name: str, dept_code: str) -> list[tuple[str, str, float]] | None:
    """Weighted (dashboard, area, weight) targets for a check: person override
    first, else the department default."""
    return PAYROLL_ATTRIBUTION_OVERRIDES.get(_norm(name)) or PROLIANT_DEPT_MAP.get(dept_code)

_SURNAME_RE = re.compile(r"^([A-Za-z][A-Za-z'.\-]*(?:\s[A-Za-z][A-Za-z'.\-]*)*),\s*$")
_DEPT_RE    = re.compile(r"Department:\s*\((\w+)\)(.+)")
_DATE_RE    = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
_TOTALS_RE  = re.compile(r"Total Earnings\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)")
_EMPID_RE   = re.compile(r"Emp Id\s+([A-Za-z0-9]+)")


# Earning / deduction / tax codes that can appear where a first name would sit.
_CODES = {
    "REG", "OT", "HOL", "VAC", "PERS", "RETRO", "SICK", "BONUS", "BRV",
    "MED", "MED-HI", "SS", "FITW", "IL", "DEN", "VIS", "HSA", "HSAF",
    "ROTH401", "VOLLIFE", "TICKET", "SALARY", "RATE", "EMP ID", "CODE",
}


def _looks_like_name(tok: str) -> bool:
    """True if `tok` reads as a person's given name rather than a payroll code."""
    t = (tok or "").strip()
    if not t or any(ch.isdigit() for ch in t):
        return False
    if t.upper() in _CODES or t.startswith("Total"):
        return False
    return bool(re.match(r"^[A-Za-z][A-Za-z'.\- ]*$", t))


def _cell(c) -> str:
    if c is None:
        return ""
    if hasattr(c, "strftime"):
        return c.strftime("%m/%d/%Y")
    return str(c)


def period_for_check_date(chk: date) -> tuple[date, date]:
    """Semi-monthly pay period a check date pays for → (start, end)."""
    if chk.day >= 15:                       # pays the 1st-15th of this month
        return date(chk.year, chk.month, 1), date(chk.year, chk.month, 15)
    # pays the 16th-end of the previous month
    y, m = (chk.year - 1, 12) if chk.month == 1 else (chk.year, chk.month - 1)
    return date(y, m, 16), date(y, m, calendar.monthrange(y, m)[1])


def business_days(start: date, end: date) -> list[date]:
    days, d = [], start
    while d <= end:
        if d.weekday() < 5:
            days.append(d)
        d += timedelta(days=1)
    return days


def parse_payroll_register(path: str | Path) -> pd.DataFrame:
    """
    One row per (employee × department × check). Columns: dept_code, dept_name,
    dashboard, area, empid, name ('LAST, FIRST'), chk_date (ISO), period_start,
    period_end (ISO), hours, gross.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    dept: tuple[str, str] | None = None
    surname = first = empid = chk = None
    saw_chk = False
    out: list[dict] = []

    for row in ws.iter_rows(values_only=True):
        cells = [_cell(c) for c in row]
        head = cells[0].strip() if cells else ""
        joined = "  ".join(c for c in cells if c)

        if head.startswith("Department:"):
            m = _DEPT_RE.search(head.split("\n")[0])
            if m and not m.group(2).strip().endswith("Total"):
                dept = (m.group(1), m.group(2).strip())
            continue

        m = _SURNAME_RE.match(head)
        if m:                                # new employee section
            surname, first, empid, chk = m.group(1), None, None, None
            saw_chk = False
            continue

        if "Chk Date" in joined:
            d = _DATE_RE.search(joined)
            if d:
                chk, saw_chk = d.group(1), True
            # The first name normally sits in this row's first cell, but some
            # blocks lead with an earning/deduction code there instead (e.g.
            # "401ER"), which would invent a second employee. Reject codes.
            if first is None and _looks_like_name(head):
                first = head

        if empid is None:
            for i, c in enumerate(cells):
                if "Emp Id" not in c:
                    continue
                mi = _EMPID_RE.search(c)
                if mi:
                    empid = mi.group(1)
                else:
                    for nxt in cells[i + 1:]:
                        if nxt.strip():
                            empid = nxt.strip()
                            break
                break

        mt = _TOTALS_RE.search(joined)
        if mt and saw_chk and surname and dept:
            chk_d = date(int(chk[6:]), int(chk[:2]), int(chk[3:5]))
            p_start, p_end = period_for_check_date(chk_d)
            out.append({
                "dept_code": dept[0], "dept_name": dept[1],
                "empid": empid,
                "name": f"{surname}, {first}" if first else surname,
                "chk_date": chk_d.isoformat(),
                "period_start": p_start.isoformat(),
                "period_end": p_end.isoformat(),
                "hours": float(mt.group(1).replace(",", "")),
                "gross": float(mt.group(2).replace(",", "")),
            })
            saw_chk = False                  # ready for this employee's next check

    df = pd.DataFrame(out)
    if df.empty:
        return df
    # Emp Id is the real identity. If one block lost its first name to a stray
    # code, heal it from the most common name that Emp Id used elsewhere.
    best = (df[df["name"].str.contains(", ")]
              .groupby("empid")["name"].agg(lambda s: s.value_counts().idxmax()))
    df["name"] = df["empid"].map(best).fillna(df["name"])

    # Attribution resolves only now: per-person overrides key on the healed name.
    tg = df.apply(lambda r: resolve_targets(r["name"], r["dept_code"]), axis=1)
    df["targets"] = tg
    df["dashboard"] = tg.apply(lambda t: t[0][0] if t else None)   # primary, for reporting
    df["area"] = tg.apply(lambda t: t[0][1] if t else None)
    df["is_split"] = tg.apply(lambda t: bool(t) and len(t) > 1)
    return df


def _norm(name: str) -> str:
    """'BARRIOS, NINI J' / 'Barrios, Nini' → 'barrios|nini' (middle initial dropped)."""
    if not name or "," not in name:
        return (name or "").strip().lower()
    last, first = name.split(",", 1)
    parts = [p for p in re.split(r"\s+", first.strip()) if p]
    if len(parts) > 1 and len(parts[-1].rstrip(".")) == 1:   # trailing middle initial
        parts = parts[:-1]
    return f"{last.strip().lower()}|{' '.join(parts).lower()}"


# Payroll name → roster tech_code, for people whose two systems disagree on the
# name itself (not just formatting) and so can never match automatically.
PAYROLL_NAME_ALIASES: dict[str, str] = {
    # Proliant drops the second surname; the roster keeps both.
    "arsovska|ruzhica brajk": "ARSORU",
}


def match_to_roster(df: pd.DataFrame, roster: list[dict]) -> pd.DataFrame:
    """Add tech_code / matched_name by normalized-name match against the roster."""
    idx = {_norm(t["name"]): t for t in roster}
    by_code = {t["tech_code"]: t for t in roster}
    df = df.copy()

    def _hit(n):
        k = _norm(n)
        if k in idx:
            return idx[k]
        code = PAYROLL_NAME_ALIASES.get(k)
        return by_code.get(code) if code else None

    hits = df["name"].apply(_hit)
    df["tech_code"] = hits.apply(lambda t: t["tech_code"] if t else None)
    df["matched_name"] = hits.apply(lambda t: t["name"] if t else None)
    df["roster_area"] = hits.apply(lambda t: t["area"] if t else None)
    df["roster_dashboard"] = hits.apply(lambda t: t["dashboard"] if t else None)
    return df


def daily_actual_rows(df: pd.DataFrame) -> list[dict]:
    """
    Explode matched, production-department payroll into daily 'actual' rows for
    labor_history. A period's gross is spread evenly across its business days —
    payroll knows the period total, not the day, so the daily figure is an
    allocation that is exact at period granularity and lets the dashboards'
    date-window sums work.
    """
    rows: list[dict] = []
    # Production-department payroll only. People on a production payroll but not
    # on the MagicTouch roster still cost the department real money, so they are
    # kept under a "PAY:" pseudo-code rather than dropped (which would understate
    # department labor %); the dashboards show that code verbatim so they stay
    # visibly distinct from roster technicians.
    ok = df[df["dashboard"].notna()].copy()
    ok["tech_code"] = ok["tech_code"].fillna("PAY:" + ok["name"])
    for _, r in ok.iterrows():
        days = business_days(date.fromisoformat(r["period_start"]),
                             date.fromisoformat(r["period_end"]))
        if not days:
            continue
        for dash, area, weight in r["targets"]:
            share = round(float(r["gross"]) * weight, 2)
            per_day = round(share / len(days), 2)
            for i, d in enumerate(days):
                # Last day carries the remainder so each share sums exactly.
                amt = (per_day if i < len(days) - 1
                       else round(share - per_day * (len(days) - 1), 2))
                rows.append({
                    "work_date": d.isoformat(), "tech_code": r["tech_code"],
                    "area": area, "dashboard": dash,
                    "pay_type": "actual", "dollars": amt,
                })
    # Same tech can hit the same slot twice (e.g. a bonus check) — sum. The key
    # includes dashboard so the two halves of a cross-department split, which
    # share an area name, stay separate.
    agg: dict[tuple, dict] = {}
    for r in rows:
        k = (r["work_date"], r["tech_code"], r["area"], r["dashboard"])
        if k in agg:
            agg[k]["dollars"] = round(agg[k]["dollars"] + r["dollars"], 2)
        else:
            agg[k] = r
    return list(agg.values())
